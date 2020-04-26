#!/usr/bin/env python3
#
#  MIT Licence
#
#  Copyright (c) 2020 Brice Rosenzweig.
#
#  Permission is hereby granted, free of charge, to any person obtaining a copy
#  of this software and associated documentation files (the "Software"), to deal
#  in the Software without restriction, including without limitation the rights
#  to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
#  copies of the Software, and to permit persons to whom the Software is
#  furnished to do so, subject to the following conditions:
#  
#  The above copyright notice and this permission notice shall be included in all
#  copies or substantial portions of the Software.
#  
#  THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#  IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#  FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#  AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#  LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
#  OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
#  SOFTWARE.
#  
#  To rebuild db form legacy:
#    ./build.py --init legacy -o out/fields.db build
#

import sqlite3
import argparse
import json
import os
import pprint
import openpyxl
from collections import defaultdict

CHANGED = 'changed'
ADDED = 'added'

def sqlite3_table_exists(conn,tablename):
    cursor = conn.execute( "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=?", (tablename,) )
    return cursor.fetchone()[0] == 1

def is_valid_language(l):
    valid = dict.fromkeys( [ 'en', 'fr', 'ja', 'de', 'it', 'es', 'pt', 'zh' ], None )
    return l in valid

def is_valid_uom_system(s):
    return s == 'metric' or s == 'statute'

def columns_ordered(cols):
    order = { 'field':10,
              'activity_type':20,
              'activity_type_id': 22,
              'parent_activity_type':30,
              'parent_activity_type_id':32,
              'category':40,
              'category_order':42,
              'field_order':50,
              'metric': 60,
              'statute': 62,
              'en': 100,
              'fr': 110,
              'de': 120,
              'it': 130,
              'es': 140,
              'pt': 150,
              'ja': 160,
              'zh': 170
              }
    return sorted( cols, key=lambda x: order[x] )

def columns_typed(cols):
    types = { 'field':'TEXT',
              'activity_type':'TEXT',
              'activity_type_id': 'INTEGER',
              'parent_activity_type':'TEXT',
              'parent_activity_type_id':'INTEGER',
              'category':'TEXT',
              'category_order':'INTEGER',
              'field_order':'INTEGER',
              'metric': 'TEXT',
              'statute': 'TEXT',
              'en': 'TEXT',
              'fr': 'TEXT',
              'de': 'TEXT',
              'it': 'TEXT',
              'es': 'TEXT',
              'pt': 'TEXT',
              'ja': 'TEXT',
              'zh': 'TEXT'
              }
    return [ '{} {}'.format( x, types[x] ) for x in columns_ordered(cols) ]
    
def dict_save_to_db(conn,listsOfDict):
    '''
    saved to db a dict of form {'table':[{'col1:'val','col2':'val'},...]}
    '''
    for table,listOfDict in listsOfDict.items():
        conn.execute( 'DROP TABLE IF EXISTS {}'.format(table) )

        sample = {}
        for one in listOfDict:
            sample.update( one )

        columns = columns_typed( sample.keys() )

        sql = 'CREATE TABLE {} ({})'.format( table,', '.join(columns) )

        conn.execute( sql )

        for type_dict in listOfDict:
            cols = ', '.join( type_dict.keys() )
            vals = ', '.join( [ ':' + x for x in type_dict.keys() ] )

            sql='INSERT INTO {} ({}) VALUES ({})'.format( table, cols, vals )
            try:
                conn.execute( sql, type_dict )
                conn.commit()
            except:
                print( 'FAILED: {} {}',sql,type_dict)
                
def dict_read_from_db(conn,tables):
    rv = {}
    for table in tables:
        if sqlite3_table_exists(conn,table):
            sql = 'SELECT * FROM {}'.format(table) ;
            res = conn.execute( sql )
            cols = res.description
            rows = []

            for row in res:
                one = {}
                for (idx,col) in enumerate(cols):
                    if row[idx] is not None:
                        one[col[0]] = row[idx]
                rows.append( one )
        rv.update( { table : rows } )
    return rv
        

def dict_save_to_excel(wb,listsOfDict):
    '''
    saved to excel a dict of form {'table':[{'col1:'val','col2':'val'},...]}
    '''
    for table,listOfDict in listsOfDict.items():
        if table in wb.sheetnames:
            wb.remove(table)
                      
        ws = wb.create_sheet(table)

        sample = {}
        for one in listOfDict:
            sample.update( one )
            
        cols = columns_ordered( sample.keys() )
        ws.append( cols )

        for type_dict in listOfDict:
            ws.append( [type_dict[x] if x in type_dict else None  for x in cols ] )

def dict_read_from_excel(wb,tables):
    rv = {}
    for table in tables:
        if table in wb.sheetnames:
            ws = wb[table]
            cells = list(ws.values)
            cols = cells[0]
            values = [dict( zip(cols,x) ) for x in cells[1:]]
            values = [ {k:v for k,v in x.items() if v is not None} for x in values ]
            rv.update( {table:values} )
    return rv
            
class ActivityType:
    def from_modern_json(input):
        rv = ActivityType()
        rv.data = {
            'activity_type' : input['typeKey'],
            'activity_type_id':input['typeId'],
            'parent_activity_type_id':input['parentTypeId'],
            }
        return rv

    def from_dict(input):
        rv = ActivityType()
        rv.data = input
        return rv
    
    def __getitem__(self,item):
        return self.data[item] if item in self.data else None

    def activityType(self):
        return self.data['activity_type']
    
    def update_parent( self, parent ):
        self.data['parent_activity_type'] = parent['activity_type']
    
    def add_display(self,language,display):
        self.data[language] = display
        
    def display(self,language):
        rv = None
        if language in self.data:
            rv = self.data[language]
        elif language == 'en':
            rv = ' '.join( [x.capitalize() for x in  self.activityType().split( '_' ) ] )
        return rv

    def languages(self):
        return [key for key in self.data.keys() if len(key) == 2]

    def update_from_dict(self,d):
        # only update the languages
        for (key,item) in d.items():
            if len(key) == 2 and item:
                self.data[key] = item
    
    def save_to_dict(self):
        return self.data
    
    def __repr__(self):
        return "ActivityType({},{},{})".format(self.data['activity_type'],self.data['parent_activity_type'], '+'.join(self.languages()) );


class ActivityTypes:
    def __init__(self):
        self.types = []
        self.verbose = False

        
    def read_from_modern_json(self,source):
        f = open(source, 'r')
        r = json.loads( f.read() )
        n = 0
        check = {}
        for item in r:
            at = ActivityType.from_modern_json(item)
            if at['activity_type'] in check:
                print( 'Duplicate type {}'.format( at ) )
            check[at['activity_type']] = at
            self.types.append( at )
            n+=1

        if self.verbose:
            print( 'add {} ActivityTypes from {}'.format( n, source ) )
        self.update_parents()

    def remap_activityType(self,atype):
        remap = {
            "snowmobiling": "snowmobiling_ws",                     
            "snow_shoe": "snow_shoe_ws",												 
            "skating": "skating_ws",
            "backcountry_skiing_snowboarding": "backcountry_skiing_snowboarding_ws",
            "skate_skiing": "skate_skiing_ws",
            "cross_country_skiing": "cross_country_skiing_ws",
            "resort_skiing_snowboarding": "resort_skiing_snowboarding_ws",
        }
        if atype in remap:
            return remap[atype]
        else:
            return atype

    def update_parents(self):
        byTypeId = {}
        for one in self.types:
            byTypeId[one['activity_type_id']] = one
        for one in self.types:
            parentId = one['parent_activity_type_id']
            if parentId:
                one.update_parent( byTypeId[ parentId ] )
        
    def activityType(self,activityType):
        for one in self.types:
            if one['activity_type'] == activityType:
                return one
        return None
                       
    def languages(self):
        rv = {}
        for one in self.types:
            rv.update( dict.fromkeys( one.languages() ) )
        return list(rv.keys())
                       
    def read_from_legacy_json(self,source,language):
        f = open(source, 'r')
        r = json.loads( f.read() )
        n = 0
        for item in r['dictionary']:
            if 'key' in item:
                n+=1
                key = self.remap_activityType( item['key'] )
                display = item['display']
                self.activityType(key).add_display(language,display)
        if self.verbose:
            print( 'added {}/{} types display for {} from {}'.format( n, len(self.types), language, source ) )
            
        self.update_parents()

    def validate_types(self):
        for one in self.types:
            display = [type_dict.display(x) for x in languages]
            if sum(1 for _ in filter( None.__ne__, display )) == 0:
                if verbose:
                    print( 'MISSING: ActivityType {}/{} has no display'.format( key, type.typeId ) )
        
    def sort(self):
        self.types = sorted( self.types, key = lambda x: (x['parent_activity_type_id'],x['activity_type_id'] ) )

    def read_from_dict(self,dicts):
        if 'gc_activity_types' in dicts:
            existings = {}
            for one in self.types:
                existings[ one['activity_type'] ] = one

            for one in dicts['gc_activity_types']:
                if one['activity_type'] in existings:
                    existings[ one['activity_type'] ].update_from_dict( one )
                else:
                    self.types.append( ActivityType.from_dict(one) )
        self.update_parents()            

    def read_from_db(self,dbname):
        conn = sqlite3.connect(dbname)
        dicts = dict_read_from_db(conn,['gc_activity_types'] )
        self.read_from_dict(dicts)
        
    def read_from_excel(self,wb):
        dicts = dict_read_from_excel(wb,['gc_activity_types'])
        self.read_from_dict(dicts)
                    
    def save_to_dict(self):
        self.sort()
        return {'gc_activity_types':[x.save_to_dict() for x in self.types]}
    
    def save_to_db(self,dbname):
        conn = sqlite3.connect(dbname)
        dicts = self.save_to_dict()
        dict_save_to_db( conn, dicts )
        
    def save_to_excel(self,wb):
        dicts = self.save_to_dict()
        dict_save_to_excel(wb,dicts)
                
    def __getitem__(self,arg):
        remap = self.remap_activityType(arg)
        alt = None
        for one in self.types:
            if one.activityType() == arg:
                return one
            if one.activityType() == remap:
                alt = one
        return alt
    
class Field:
    '''
    units: array of Dict{ 'activityType':xxx,'metrics':xxx,'statute':xxx }
    fieldDisplayNameByLanguage Dict(language:display)
    order: array of {'category':STR,'field_order':NUMBER, ['activityType':STR or NULL] }
    changes: array of dict {'what':xxx, 'old':xxx,'new':xxx }
    '''    
    def __init__(self, key):
        self.key = key
        self.fieldDisplayNameByLanguage = dict()
        self.units = []
        self.orders = []

        self.changes = []

    def sort_key(self):
        return self.key

    def reset_changes(self):
        self.changes = []
        
    def report_changes(self):
        for change in self.changes:
            print( '{} changed {} from {} to {}'.format( self.key, change['what'], change['old'], change['new'] ) )
    
    def add_display(self, language, displayName ):
        rv = None
        if not displayName:
            return rv
        
        if language not in self.fieldDisplayNameByLanguage:
            if displayName != self.key:
                rv = ADDED
                self.fieldDisplayNameByLanguage[language] = displayName
        else:
            if displayName != self.key and displayName != self.fieldDisplayNameByLanguage[language]:
                rv = CHANGED
                self.changes.append( {'what':language,'old':self.fieldDisplayNameByLanguage[language],'new':displayName} )
                self.fieldDisplayNameByLanguage[language] = displayName
        return rv
                
    def add_uom(self, system, uom, activityType ):
        rv = None
        found = None
        replace = False
        for one in self.units:
            if one['activity_type'] == activityType:
                found = one
                break
            
        if found is None:
            found = { 'field':self.key,'activity_type':activityType}
            self.units.append(found)
            rv = ADDED
        if system in found and found[system] != uom:
            replace = True
            self.changes.append( {'what':(activityType,system),'old':found[system],'new':uom} )
            rv = CHANGED
            
        found[system] = uom

        return rv

    def add_category_and_order(self, info):
        '''
        expected info = {'category':STR,'field_order':NUMBER, ['activityType':STR or NULL] }
        if no activityType or null applies to all
        '''
        rv = None
        
        existing = None
        if 'activity_type' in info and info['activity_type']:
            for one in self.orders:
                if 'activity_type' in one and one['activity_type'] == info['activity_type']:
                    existing = one
        else:
            for one in self.orders:
                if 'activity_type' not in one:
                    existing = one

        if existing:
            if existing != one:
                self.changes.append( { 'what':'order','old':dict(existing),'new':info} )
                existing.update(one)
                rv = CHANGED
        else:
            self.orders.append( info )
            rv = ADDED

        return rv
        
    def simplify_uom(self):
        all = None
        for one in self.units:
            if one['activity_type'] == 'all':
                all = one
                break
        simplified = []
        if all:
            simplified.append( all )
            for one in self.units:
                for k,v in one.items():
                    if k!= 'activity_type' and k in all and all[k] != v:
                        simplified.append(one)
                        break
            self.units = simplified
        else:
            print( 'missing all {}'.format( self.units ) )
            
    def display_dicts(self):
        rv = {'field':self.key }
        rv.update( self.fieldDisplayNameByLanguage )
        return [ rv ]

    def uom_dicts(self):
        return self.units

    def order_dicts(self):
        rv = []
        rv.extend( self.orders )
        return rv
    
    def __repr__(self):
        return "Field('{}',{},{})".format(self.key, pprint.pformat( self.fieldDisplayNameByLanguage ), pprint.pformat( self.units ) );

class Fields:
    '''
    fields is a dict of {'key':Field(...),...}
    activityTypes = ActivityTypes()
    '''
    def __init__(self,types):
        self.fields = dict()
        self.activityTypes = types
        self.verbose = types.verbose
        
        self.changes = defaultdict(list)
        self.adds = defaultdict(list)
        self.checks = defaultdict(list)

    def field(self,key):
        if key not in self.fields:
            self.fields[key] = Field(key)

        return self.fields[key]

    def read_from_legacy_db(self, db_from, language, unitsystem, table = 'gc_fields' ):
        conn = sqlite3.connect(db_from)
        cursor = conn.execute('select * from {}'.format(table))
        
        n = {}
        n_new = {}
        
        for row in cursor:
            (fieldKey, activityType, displayName, uom ) = row
            if self.activityTypes[fieldKey]:
                continue
            
            if fieldKey not in self.fields:
                n_new[fieldKey] = 1
                self.fields[fieldKey] = Field(fieldKey)

            r1 = self.fields[fieldKey].add_display( language, displayName )
            r2 = self.fields[fieldKey].add_uom( unitsystem, uom, activityType )
            if r1 or r2:
                n[fieldKey] = 1

        for field in self.fields.values():
            field.simplify_uom()
        if self.verbose:
            print( 'added {}/{} field display (new {}) for {} from {}'.format( len(n), len( self.fields ), len(n_new), language, db_from ) )
            
    def fix_legacy_unit_system_missing(self):
        metric_to_statute = dict()
        # collect known conversions
        for key,field in self.fields.items():
            for one in field.units:
                if 'metric' in one and 'statute' in one and one['metric'] != one['statute']:
                    metric_to_statute[ one['metric'] ] = one['statute']

        for key,field in self.fields.items():
            for one in field.units:
                u = one['metric']
                t = one['activity_type']
                n_u = None
                if u in metric_to_statute:
                    n_u = metric_to_statute[u]
                    if n_u == 'foot' and 'Elevation' not in key:
                        n_u = 'yard'
                if n_u:
                    if 'statute' not in one:
                        if self.verbose:
                            print( 'updating missing {}/{} to {} (from {})'.format( key, t, n_u, u) )
                        field.add_uom( 'statute',n_u, t)
                    if  n_u != one['statute']:
                        if self.verbose:
                            print( 'fixing inconsistent {}/{} {} != {} (from {})'.format( key, t, one['statute'], n_u, u) )
                        field.add_uom( 'statute',n_u, t)


                if 'Elevation' in key and one['metric'] != 'meter':
                    if self.verbose:
                        print( 'fixing elevation {}/{} {}/{} to {}/{}'.format( key, t, one['metric'], one['statute'] if t in one['statute'] else None , 'meter', 'foot') )

                    field.add_uom( 'metric','meter', t)
                    field.add_uom( 'statute','foot', t)

    def find(self,needle):
        rv = []
        for (key,field) in self.fields.items():
            if needle in key:
                rv.append( field )

        return rv

    def reset_status(self,table=None):
        keys = [table] if table else list(set().union(self.changes.keys(),self.adds.keys(),self.checks.keys()))
        for k in keys:
            self.adds[k] = []
            self.changes[k] = []
            self.checks[k] = []

    def record_status(self,table,field,field_status=None):
        self.checks[table].append( field )
        
        if field_status == CHANGED:
            self.changes[table].append( field )
        elif field_status == ADDED:
            self.adds[table].append( field )

    def report_changes(self):
        for one in self.types:
            one.report_changes()
            
    def report_status(self,table=None):
        keys = [table] if table else list(set().union(self.changes.keys(),self.adds.keys()))
        for k in keys:
            print( '{}: read {} added {} changed {}'.format( k, len(self.checks[k]) if k in self.checks else 0, len(self.adds[k]) if k in self.adds else 0, len(self.changes[k]) if k in self.changes else 0) )

    def report_diffs(self,table=None):
        keys = [table] if table else list(set().union(self.changes.keys(),self.adds.keys()))
        for k in keys:
            print( '{}: read {} added {} changed {}'.format( k, len(self.checks[k]) if k in self.checks else 0, len(self.adds[k]) if k in self.adds else 0, len(self.changes[k]) if k in self.changes else 0) )
            if k in self.adds:
                for one in self.adds[k]:
                    print( 'add {}'.format( one.key ) )
            if k in self.changes:
                for one in self.changes[k]:
                    one.report_changes()
            
    def read_field_order_from_dict(self,d):
        if 'gc_fields_order' in d:
            order = d['gc_fields_order']
            self.reset_status('gc_fields_order')
            for one in order:
                field = self.field(one['field'])
                rv = field.add_category_and_order(one)
                self.record_status('gc_fields_order',field,rv)
            if self.verbose:
                self.report_status('gc_fields_order')

    def read_field_display_from_dict(self,d):
        if 'gc_fields_display' in d:
            self.reset_status('gc_fields_display')
            display = d['gc_fields_display']
            for one in display:
                field = self.field( one['field'] )
                field_status = None
                for key,val in one.items():
                    if is_valid_language(key):
                        rv = field.add_display( key,val )
                        if rv and not field_status:
                            field_status = rv
                self.record_status('gc_fields_display',field,field_status)
            if self.verbose:
                self.report_status('gc_fields_display')

    def read_field_uom_from_dict(self,d):
        if 'gc_fields_uom' in d:
            self.reset_status('gc_fields_uom')
            uom = d['gc_fields_uom']
            for one in uom:
                field = self.field(one['field'])
                field_status = None
                for (k,v) in one.items():
                    if is_valid_uom_system(k):
                        rv = field.add_uom(k,v,one['activity_type'])
                        if rv and not field_status:
                            field_status = rv
                        
                self.record_status('gc_fields_uom',field,field_status)
            if self.verbose:
                self.report_status('gc_fields_uom')
    def read_from_dict(self,dicts):
        self.read_field_order_from_dict( dicts )
        self.read_field_display_from_dict( dicts )
        self.read_field_uom_from_dict(dicts)
        
    def read_from_db(self,dbname):
        conn = sqlite3.connect(dbname)
        dicts = dict_read_from_db(conn,['gc_fields_order','gc_fields_display','gc_fields_uom'] )
        self.read_from_dict(dicts)
        
    def read_from_excel(self,wb):
        dicts = dict_read_from_excel(wb,['gc_fields_order','gc_fields_display','gc_fields_uom'])    
        self.read_from_dict(dicts)
        
    def save_to_db(self,dbname):
        dicts = self.save_to_dict()
        conn = sqlite3.connect(dbname)
        dict_save_to_db(conn,dicts)

    def save_to_excel(self,wb):
        dicts = self.save_to_dict()
        dict_save_to_excel(wb,dicts)
        
    def save_to_dict(self):
        display = []
        uom = []
        order = []
        keys = sorted( self.fields.keys(), key = lambda x: self.fields[x].sort_key() )

        for key in keys:
            field = self.fields[key]
            display.extend( field.display_dicts() )
            uom.extend( field.uom_dicts() )
            order.extend( field.order_dicts() )
            
        return { 'gc_fields_uom': uom, 'gc_fields_display': display, 'gc_fields_order': order }
    
    
class Category :
    def __init__(self,name,category_order,display_name):
        self.name = name
        self.category_order = category_order
        self.categoryDisplayNameByLanguage = {'en':display_name}

    def save_to_dict(self):
        rv = {'category':self.name,'category_order':self.category_order}
        rv.update( self.categoryDisplayNameByLanguage )
        return rv

class Categories :
    def __init__(self):
        self.categories = {}
        self.verbose = False

    def add_category(self,category):
        self.categories[ category.name ] = category

    def read_from_dict(self,dicts):
        if 'gc_category_order' in dicts:
            start = len(self.categories)
            for one in dicts['gc_category_order']:
                if len(one):
                    cat = Category(one['category'],one['category_order'],one['en'])
                    self.add_category( cat )
            if self.verbose:
                print( 'gc_category_order: add {} (was {}))'.format( len( self.categories ), start ) )

    def read_from_db(self,dbname):
        conn = sqlite3.connect(dbname)
        dicts = dict_read_from_db(conn,['gc_category_order'])
        self.read_from_dict(dicts)

    def read_from_excel(self,wb):
        dicts = dict_read_from_excel(wb,['gc_category_order'])
        self.read_from_dict(dicts)

    def save_to_db(self,dbname):
        conn = sqlite3.connect(dbname)
        dicts = self.save_to_dict()
        dict_save_to_db(conn,dicts)

    def save_to_excel(self,wb):
        dicts = self.save_to_dict()
        dict_save_to_excel(wb,dicts)
        
    def save_to_dict(self):
        rv = []
        ordered = sorted(self.categories.keys(), key=lambda x: self.categories[x].category_order )
        for key in ordered:
            rv.append( self.categories[ key ].save_to_dict() )
        return {'gc_category_order':rv}
    
class Driver :
    def __init__(self,args):
        self.args = args
        self.verbose = args.verbose

    def init_latest(self):
        base = self.args.base
        self.types = ActivityTypes()
        self.types.verbose = self.verbose
        self.types.read_from_modern_json( 'download/activity_types_modern.json' )
        if self.verbose:
            print( 'read {}'.format( base ) )
        self.types.read_from_db(base)
        self.fields = Fields(self.types)
        self.fields.verbose = self.verbose
        self.fields.read_from_db(base)
        self.categories = Categories()
        self.categories.verbose = self.verbose
        self.categories.read_from_db( base )
        
    def init_legacy(self):
        self.types = ActivityTypes()
        self.types.verbose = self.verbose
        self.types.read_from_modern_json( 'download/activity_types_modern.json' )
        self.fields = Fields(self.types)
        self.fields.verbose = self.verbose
        
        self.languages = [ 'en', 'fr', 'ja', 'de', 'it', 'es', 'pt', 'zh' ]

        for lang in self.languages:
            self.types.read_from_legacy_json( 'cached/activity_types_{}.json'.format( lang ), lang )
        for lang in self.languages:
            self.fields.read_from_legacy_db( 'cached/fields_{}_metric.db'.format( lang ), lang, 'metric' )
        self.fields.read_from_legacy_db( 'cached/fields_en_statute.db', 'en', 'statute' )
        self.fields.read_from_excel(openpyxl.load_workbook(filename='edit/gc_fields_manual.xlsx'))
        self.fields.fix_legacy_unit_system_missing()

        self.categories = Categories()
        self.categories.verbose = self.verbose
        self.categories.read_from_db( 'edit/fields_order.db' )
        self.fields.read_from_db( 'edit/fields_order.db' )

    def init_empty(self):
        self.types = ActivityTypes()
        self.types.verbose = self.args.verbose
        self.types.read_from_modern_json( 'download/activity_types_modern.json' )
        self.fields = Fields(types)
        self.fields.verbose = self.verbose
        self.categories = Categories()
        self.categories.verbose = self.verbose
        
    def cmd_build(self):
        for fn in self.args.files:
            self.process_file( fn )

        output = self.args.output
        if output.endswith( '.db' ):
            self.fields.save_to_db(self.args.output )
            self.types.save_to_db( self.args.output )
            self.categories.save_to_db(self.args.output )
        if output.endswith( '.json' ):
            f = self.fields.save_to_dict()
            a = self.types.save_to_dict( )
            f.update( a )
            with open(output, 'w') as of:
                json.dump(f, of, indent=2, sort_keys=True)
        if output.endswith( '.xlsx' ):
            wb = openpyxl.Workbook()
            self.fields.save_to_excel(wb)
            self.types.save_to_excel(wb)
            self.categories.save_to_excel(wb)
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            wb.save( output )
            

    def cmd_diff(self):
        for fn in self.args.files:
            self.process_file( fn )
            self.fields.report_diffs()

    def load_db(self,dbname):
        if os.path.exists( dbname ):
            conn = sqlite3.connect( dbname )
            self.types.read_from_db(conn)
            self.categories.read_from_db(conn)
            self.fields.read_from_db(conn)

    def load_excel(self,xlname):
        if os.path.exists( xlname ):
            wb = openpyxl.load_workbook(filename=xlname)
            self.types.read_from_excel(wb)
            self.fields.read_from_excel(wb)
            self.categories.read_from_excel(wb)


    def load_from_json(self,jname):
        with open(jname,'r') as jf:
            jd = json.load(jf )
            self.types.read_from_dict(jd)
            self.fields.read_from_dict(jd)
            self.categories.read_from_dict(jd)
            
    def process_file(self,fn):
        if os.path.exists( fn ):
            if self.verbose:
                print( 'processing {}'.format( fn ) )
            if fn.endswith( '.db' ) :
                self.load_db( fn )
            elif fn.endswith( '.json' ):
                self.load_json( fn )
            elif fn.endswith( '.xlsx' ):
                self.load_excel( fn )
                
if __name__ == "__main__":
                
    commands = {
        'build':{'attr':'cmd_build','help':'Rebuild database'},
        'diff':{'attr':'cmd_diff','help':'Show Diff if rebuild'},
    }

    init = {
        'legacy':{'attr':'init_legacy','help':'Build from legacy files'},
        'empty':{'attr':'init_empty','help':'Build empty base'},
        'latest':{'attr':'init_latest', 'help':'Build with latest file'},
    }
    
    description = "\n".join( [ '  {}: {}'.format( k,v['help'] ) for (k,v) in commands.items() ] )
    init_desc = "\n".join( [ '  {}: {}'.format( k,v['help'] ) for (k,v) in init.items() ] )

    languages = [ 'en', 'fr', 'ja', 'de', 'it', 'es', 'pt', 'zh' ]
    what = ['activity_type','unit','display']
    
    parser = argparse.ArgumentParser( description='Check configuration', formatter_class=argparse.RawTextHelpFormatter )
    parser.add_argument( 'command', metavar='Command', help='command to execute:\n' + description)
    parser.add_argument( '-b', '--base', default='out/fields.db', help='db file to init latest from' )
    parser.add_argument( '-s', '--save', action='store_true', help='save output otherwise just print' )
    parser.add_argument( '-o', '--output', help='output file' )
    parser.add_argument( '-i', '--init', help='init method (default legacy)\n' + init_desc, default='latest' )
    parser.add_argument( '-v', '--verbose', action='store_true', help='verbose output' )
    parser.add_argument( '-w', '--what', help='list what to show or process, defaults to {}'.format( '+'.join(what)), default='+'.join(what))
    parser.add_argument( '-l', '--languages', help='list of languages to show or process. defaults to {}'.format( '+'.join( languages ) ) , default='+'.join(languages))
    parser.add_argument( 'files',    metavar='FILES', nargs='*', help='files to process' )
    args = parser.parse_args()

    command = Driver(args)

    if args.init in init:
        getattr(command,init[args.init]['attr'])()
        if args.command in commands:
            getattr(command,commands[args.command]['attr'])()
        else:
            print( 'Invalid command "{}"'.format( args.command) )
            parser.print_help()
    else:
        print( 'Invalid init "{}"'.format( args.command) )
        parser.print_help()

        
        
